#!/usr/bin/env python3
"""
Microsoft 365 Tenant Discovery Report — Self-Hosted Web App
Upload Fly Migration Excel reports → interactive dashboard → PDF export.
Supports: Exchange Online, Microsoft 365 Groups, Teams, OneDrive, SharePoint Online
"""

import os, io, uuid, json
from datetime import datetime
import openpyxl
from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, send_file, session, jsonify
)
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.backends.backend_pdf import PdfPages

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "change-me-in-production")
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = 32 * 1024 * 1024

# ─── Utility ──────────────────────────────────────────────────────────────────

def safe_int(v):
    if v is None: return 0
    try: return int(v)
    except: return 0

def safe_float(v):
    if v is None: return 0.0
    v = str(v).strip()
    if v.startswith("<"): return 0.0
    try: return float(v)
    except: return 0.0

def find_col(headers, *keyword_sets):
    """Find column index matching keywords. keyword_sets are OR groups, each is an AND list."""
    for keywords in keyword_sets:
        for i, h in enumerate(headers):
            if all(k in h for k in keywords):
                return i
    return None

def classify_activity(date_str, now=None):
    """Classify activity into time buckets."""
    if now is None: now = datetime.now()
    date_str = str(date_str).strip()
    if date_str in ("-", "", "None", " -"):
        return "inactive"
    try:
        parts = date_str.split("/")
        d = datetime(int(parts[2]), int(parts[0]), int(parts[1]))
        days = (now - d).days
        if days <= 180: return "within_6m"
        elif days <= 365: return "6m_to_1y"
        else: return "older_1y"
    except:
        return "inactive"

def is_inactive(date_str, days_threshold=30, now=None):
    if now is None: now = datetime.now()
    date_str = str(date_str).strip()
    if date_str in ("-", "", "None", " -"): return True
    try:
        parts = date_str.split("/")
        d = datetime(int(parts[2]), int(parts[0]), int(parts[1]))
        return (now - d).days > days_threshold
    except:
        return True


# ─── Parsers ──────────────────────────────────────────────────────────────────

def parse_exchange(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]

    def fc(*kw_sets): return find_col(headers, *kw_sets)

    c_email = fc(["email"]) or 0
    c_type = fc(["type"]) or 1
    c_items = None
    for i,h in enumerate(headers):
        if "items" in h and "archive" not in h and "deleted" not in h:
            c_items = i; break
    c_storage = None
    for i,h in enumerate(headers):
        if "storage" in h and "archive" not in h:
            c_storage = i; break
    c_deleted = None
    for i,h in enumerate(headers):
        if "deleted" in h and "items" in h and "archive" not in h and "size" not in h:
            c_deleted = i; break
    c_del_size = None
    for i,h in enumerate(headers):
        if "deleted" in h and "size" in h and "archive" not in h:
            c_del_size = i; break
    c_last = fc(["last", "activity"], ["last"])
    c_archive = fc(["archive", "status"], ["archive"])

    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[c_email]: continue
        rows.append({
            "email": str(row[c_email]).strip(),
            "type": str(row[c_type]).strip() if c_type is not None else "Unknown",
            "items": safe_int(row[c_items] if c_items is not None else 0),
            "storage": safe_float(row[c_storage] if c_storage is not None else 0),
            "deleted": safe_int(row[c_deleted] if c_deleted is not None else 0),
            "del_size": safe_float(row[c_del_size] if c_del_size is not None else 0),
            "last_activity": str(row[c_last]).strip() if c_last is not None and row[c_last] else "-",
            "archive": str(row[c_archive]).strip() if c_archive is not None and row[c_archive] else "Disabled",
        })
    return rows

def parse_groups(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]: continue
        rows.append({
            "email": str(row[find_col(headers, ["email"]) or 0]).strip(),
            "name": str(row[find_col(headers, ["group name"], ["name"]) or 1]).strip(),
            "privacy": str(row[find_col(headers, ["privacy"]) or 2]).strip(),
            "owners": safe_int(row[find_col(headers, ["owners"]) or 3]),
            "members": safe_int(row[find_col(headers, ["members"]) or 4]),
            "guests": safe_int(row[find_col(headers, ["guests"]) or 5]),
            "teams_status": str(row[find_col(headers, ["teams"]) or 6]).strip(),
            "last_activity": str(row[find_col(headers, ["last", "activity"]) or 7]).strip(),
            "site_title": str(row[find_col(headers, ["site", "title"]) or 8]).strip(),
            "site_size": safe_float(row[find_col(headers, ["team site size"], ["site size"]) or 10]),
            "mailbox_size": safe_float(row[find_col(headers, ["mailbox size"]) or 11]),
            "total_size": safe_float(row[find_col(headers, ["total size"]) or 12]),
        })
    return rows

def parse_teams(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]: continue
        rows.append({
            "name": str(row[find_col(headers, ["name"]) or 0]).strip(),
            "email": str(row[find_col(headers, ["email"]) or 1]).strip(),
            "privacy": str(row[find_col(headers, ["privacy"]) or 2]).strip(),
            "owners": safe_int(row[find_col(headers, ["owners"]) or 3]),
            "members": safe_int(row[find_col(headers, ["members"]) or 4]),
            "guests": safe_int(row[find_col(headers, ["guests"]) or 5]),
            "total_channels": safe_int(row[find_col(headers, ["total channel"]) or 6]),
            "private_channels": safe_int(row[find_col(headers, ["private channel"]) or 7]),
            "standard_channels": safe_int(row[find_col(headers, ["standard channel"]) or 8]),
            "shared_channels": safe_int(row[find_col(headers, ["shared channel"]) or 9]),
            "status": str(row[find_col(headers, ["status"]) or 10]).strip(),
            "created": str(row[find_col(headers, ["created"]) or 11]).strip(),
            "last_activity": str(row[find_col(headers, ["last", "activity"]) or 12]).strip(),
            "site_size": safe_float(row[find_col(headers, ["team site size"]) or 15]),
            "channel_site_size": safe_float(row[find_col(headers, ["channel site"]) or 16]),
            "mailbox_size": safe_float(row[find_col(headers, ["mailbox"]) or 17]),
            "total_size": safe_float(row[find_col(headers, ["total size"]) or 18]),
        })
    return rows

def parse_onedrive(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]: continue
        rows.append({
            "url": str(row[find_col(headers, ["url"]) or 0]).strip(),
            "upn": str(row[find_col(headers, ["principal"], ["user principal"]) or 1]).strip(),
            "display_name": str(row[find_col(headers, ["display"]) or 2]).strip(),
            "storage": safe_float(row[find_col(headers, ["storage"]) or 3]),
            "files": safe_int(row[find_col(headers, ["files"]) or 4]),
            "last_activity": str(row[find_col(headers, ["last", "activity"]) or 5]).strip(),
        })
    return rows

def parse_sharepoint(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]
    headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]: continue
        rows.append({
            "name": str(row[find_col(headers, ["site name"], ["name"]) or 0]).strip(),
            "url": str(row[find_col(headers, ["url"]) or 1]).strip(),
            "template": str(row[find_col(headers, ["template"]) or 2]).strip(),
            "storage": safe_float(row[find_col(headers, ["storage"]) or 3]),
            "files": safe_int(row[find_col(headers, ["files"]) or 4]),
            "last_activity": str(row[find_col(headers, ["last", "activity"]) or 5]).strip(),
        })
    return rows


# ─── Statistics computation ───────────────────────────────────────────────────

def compute_exchange_stats(rows):
    tc = {}
    for r in rows: tc[r["type"]] = tc.get(r["type"], 0) + 1
    storage_buckets_def = [("0-10 GB",0,10),("10-20 GB",10,20),("20-30 GB",20,30),("30-40 GB",30,40),("40-50 GB",40,50),("> 50 GB",50,float("inf"))]
    sb = [{"label":l,"count":sum(1 for r in rows if lo<=r["storage"]<hi)} for l,lo,hi in storage_buckets_def]
    ae = sum(1 for r in rows if r["archive"].lower()=="enabled")
    return {
        "total_mailboxes": len(rows), "total_items": sum(r["items"] for r in rows),
        "total_storage": round(sum(r["storage"] for r in rows),2),
        "total_deleted": sum(r["deleted"] for r in rows),
        "total_del_size": round(sum(r["del_size"] for r in rows),2),
        "inactive": sum(1 for r in rows if is_inactive(r["last_activity"])),
        "type_counts": tc, "storage_buckets": sb,
        "archive_enabled": ae, "archive_disabled": len(rows)-ae,
        "top_storage": sorted(rows, key=lambda r:r["storage"], reverse=True)[:10],
        "top_items": sorted(rows, key=lambda r:r["items"], reverse=True)[:10],
        "rows": rows,
    }

def compute_groups_stats(rows):
    privacy = {}
    for r in rows: privacy[r["privacy"]] = privacy.get(r["privacy"],0) + 1
    no_owners = sum(1 for r in rows if r["owners"]==0)
    inactive = sum(1 for r in rows if is_inactive(r["last_activity"]))
    guests = sum(1 for r in rows if r["guests"]>0)
    sb_def = [("0-10 GB",0,10),("10-100 GB",10,100),("100-500 GB",100,500),("500 GB-1 TB",500,1000),("> 1 TB",1000,float("inf"))]
    sb = [{"label":l,"count":sum(1 for r in rows if lo<=r["site_size"]<hi)} for l,lo,hi in sb_def]
    return {
        "total": len(rows), "no_owners": no_owners, "inactive": inactive,
        "with_guests": guests, "privacy": privacy, "storage_buckets": sb, "rows": rows,
    }

def compute_teams_stats(rows):
    privacy = {}
    for r in rows: privacy[r["privacy"]] = privacy.get(r["privacy"],0)+1
    total_ch = sum(r["total_channels"] for r in rows)
    priv_ch = sum(r["private_channels"] for r in rows)
    std_ch = sum(r["standard_channels"] for r in rows)
    shared_ch = sum(r["shared_channels"] for r in rows)
    no_owners = sum(1 for r in rows if r["owners"]==0)
    inactive = sum(1 for r in rows if is_inactive(r["last_activity"]))
    guests = sum(1 for r in rows if r["guests"]>0)
    sb_def = [("0-10 GB",0,10),("10-100 GB",10,100),("100-500 GB",100,500),("500 GB-1 TB",500,1000),("> 1 TB",1000,float("inf"))]
    sb = [{"label":l,"count":sum(1 for r in rows if lo<=r["site_size"]<hi)} for l,lo,hi in sb_def]
    return {
        "total": len(rows), "no_owners": no_owners, "inactive": inactive,
        "with_guests": guests, "total_channels": total_ch,
        "channel_types": {"Standard channel": std_ch, "Private channel": priv_ch, "Shared channel": shared_ch},
        "privacy": privacy, "storage_buckets": sb, "rows": rows,
    }

def compute_onedrive_stats(rows):
    sb_def = [("0-10 GB",0,10),("10-20 GB",10,20),("20-30 GB",20,30),("30-40 GB",30,40),("40-50 GB",40,50),("> 50 GB",50,float("inf"))]
    sb = [{"label":l,"count":sum(1 for r in rows if lo<=r["storage"]<hi)} for l,lo,hi in sb_def]
    return {
        "total": len(rows),
        "total_storage": round(sum(r["storage"] for r in rows),2),
        "total_files": sum(r["files"] for r in rows),
        "storage_buckets": sb,
        "top_files": sorted(rows, key=lambda r:r["files"], reverse=True)[:10],
        "top_storage": sorted(rows, key=lambda r:r["storage"], reverse=True)[:10],
        "rows": rows,
    }

def compute_sharepoint_stats(rows):
    sb_def = [("0-10 GB",0,10),("10-100 GB",10,100),("100-500 GB",100,500),("500 GB-1 TB",500,1000),("> 1 TB",1000,float("inf"))]
    sb = [{"label":l,"count":sum(1 for r in rows if lo<=r["storage"]<hi)} for l,lo,hi in sb_def]
    templates = {}
    for r in rows: templates[r["template"]] = templates.get(r["template"],0)+1
    activity = {"Within 6 months":0, "6 months to 1 year":0, "Older than 1 year":0}
    active_count = 0
    for r in rows:
        c = classify_activity(r["last_activity"])
        if c == "within_6m": activity["Within 6 months"]+=1; active_count+=1
        elif c == "6m_to_1y": activity["6 months to 1 year"]+=1; active_count+=1
        elif c == "older_1y": activity["Older than 1 year"]+=1; active_count+=1
    return {
        "total": len(rows),
        "total_storage": round(sum(r["storage"] for r in rows),2),
        "total_files": sum(r["files"] for r in rows),
        "storage_buckets": sb, "templates": templates, "activity": activity,
        "top_files": sorted(rows, key=lambda r:r["files"], reverse=True)[:10],
        "top_storage": sorted(rows, key=lambda r:r["storage"], reverse=True)[:10],
        "rows": rows,
    }


# ─── PDF generation ──────────────────────────────────────────────────────────

PRIMARY="#2563eb"; SECONDARY="#f59e0b"; TERTIARY="#7c3aed"; TEAL="#0d9488"
ACCENT="#10b981"; TEXT="#1e293b"; TEXT_SEC="#64748b"; BG="#f0f2f5"
CARD="#ffffff"; BORDER="#e2e8f0"
DONUT_COLORS=["#2563eb","#f59e0b","#7c3aed","#10b981","#ef4444","#8b5cf6"]

def _setup_plt():
    plt.rcParams.update({
        "font.family":"DejaVu Sans","font.size":9,"axes.facecolor":CARD,
        "figure.facecolor":BG,"text.color":TEXT,"axes.labelcolor":TEXT_SEC,
        "xtick.color":TEXT_SEC,"ytick.color":TEXT_SEC,
    })

def _draw_stat(ax, value, label):
    ax.set_xlim(0,1); ax.set_ylim(0,1); ax.set_facecolor(CARD)
    for s in ax.spines.values(): s.set_edgecolor(BORDER); s.set_linewidth(0.8)
    ax.set_xticks([]); ax.set_yticks([])
    ax.text(0.5,0.6,str(value),ha="center",va="center",fontsize=16,fontweight="bold",color=TEXT)
    ax.text(0.5,0.22,label,ha="center",va="center",fontsize=6.5,fontweight="medium",color=TEXT_SEC)

def _draw_donut(ax, sizes, colors, total, center_label):
    if sum(sizes)==0: sizes=[1]; colors=["#e2e8f0"]
    ax.pie(sizes,colors=colors[:len(sizes)],startangle=90,wedgeprops=dict(width=0.38,edgecolor="white",linewidth=2))
    ax.text(0,0.05,str(total),ha="center",va="center",fontsize=20,fontweight="bold",color=TEXT)
    ax.text(0,-0.18,center_label,ha="center",va="center",fontsize=9,fontweight="medium",color=PRIMARY)
    ax.set_aspect("equal")

def _draw_legend(ax, labels, sizes, colors):
    ax.axis("off"); ax.set_xlim(0,1); ax.set_ylim(0,1)
    n = len(labels)
    for i,(lab,sz) in enumerate(zip(labels,sizes)):
        y = 0.85 - i*(0.7/max(n-1,1)) if n>1 else 0.5
        rect = mpatches.FancyBboxPatch((0.02,y-0.04),0.06,0.06,boxstyle="round,pad=0.01",
                                        facecolor=colors[i%len(colors)],edgecolor="none")
        ax.add_patch(rect)
        ax.text(0.14,y,lab,va="center",fontsize=8.5,color=TEXT)
        ax.text(0.92,y,str(sz),va="center",ha="right",fontsize=9.5,fontweight="bold",color=TEXT)

def _draw_bar(ax, labels, values, color=TEAL):
    bars = ax.bar(labels, values, color=color, width=0.55, edgecolor="white", linewidth=0.5, zorder=3)
    ax.set_facecolor(CARD)
    ax.spines["top"].set_visible(False); ax.spines["right"].set_visible(False)
    ax.spines["left"].set_visible(False); ax.spines["bottom"].set_edgecolor(BORDER)
    ax.tick_params(axis="x",length=0); ax.yaxis.grid(True,linestyle="--",alpha=0.4,color="#ccc")
    ax.set_axisbelow(True)
    for b,v in zip(bars,values):
        if v>0: ax.text(b.get_x()+b.get_width()/2,b.get_height()+max(max(values)*0.02,0.1),str(v),ha="center",va="bottom",fontsize=8,fontweight="bold",color=TEXT)

def _draw_hbar(ax, labels, values, color=PRIMARY, fmt=None):
    bars = ax.barh(range(len(labels)),values,color=color,height=0.6,alpha=0.85,edgecolor="white",linewidth=0.5,zorder=3)
    ax.set_yticks(list(range(len(labels)))); ax.set_yticklabels(labels,fontsize=7.5)
    ax.set_facecolor(CARD)
    for s in ["top","right","bottom","left"]: ax.spines[s].set_visible(False)
    ax.xaxis.grid(True,linestyle="--",alpha=0.3,color="#ccc"); ax.set_axisbelow(True)
    ax.tick_params(axis="y",length=0); ax.tick_params(axis="x",length=0)
    for b,v in zip(bars,values):
        txt = fmt(v) if fmt else str(v)
        ax.text(b.get_width()+max(max(values)*0.02,0.3),b.get_y()+b.get_height()/2,txt,va="center",fontsize=7,fontweight="bold",color=TEXT)

def _draw_table(ax, col_labels, table_data, col_widths=None):
    ax.axis("off")
    tbl = ax.table(cellText=table_data,colLabels=col_labels,loc="upper center",cellLoc="center",colColours=[PRIMARY]*len(col_labels))
    tbl.auto_set_font_size(False); tbl.set_fontsize(5.8); tbl.scale(1,1.12)
    for (ri,ci),cell in tbl.get_celld().items():
        cell.set_edgecolor(BORDER); cell.set_linewidth(0.5)
        if col_widths and ci<len(col_widths): cell.set_width(col_widths[ci])
        if ri==0:
            cell.set_text_props(color="white",fontweight="bold",fontsize=6); cell.set_facecolor(PRIMARY); cell.set_height(0.026)
        else:
            cell.set_facecolor("#fafbfc" if ri%2==0 else CARD); cell.set_text_props(color=TEXT,fontsize=5.5); cell.set_height(0.024)


def _pdf_exchange(pdf, stats, detailed):
    fig = plt.figure(figsize=(14,9.5)); fig.patch.set_facecolor(BG)
    fig.text(0.04,0.96,"Exchange Online — Tenant Overview",fontsize=17,fontweight="bold",color=TEXT,va="top")
    # Stats
    sd = [(str(stats["total_mailboxes"]),"Total Mailboxes"),(f'{stats["total_items"]:,}',"Total Items"),
          (f'{stats["total_storage"]:.2f}',"Storage (GB)"),(f'{stats["total_deleted"]:,}',"Deleted Items"),
          (f'{stats["total_del_size"]:.2f}',"Del. Size (GB)"),(str(stats["inactive"]),"Inactive")]
    for i,(v,l) in enumerate(sd):
        ax=fig.add_axes([0.04+i*0.155,0.84,0.14,0.07]); _draw_stat(ax,v,l)
    # Type donut
    fig.text(0.04,0.80,"Mailboxes by type",fontsize=11,fontweight="bold",color=PRIMARY)
    ordered = ["User mailbox","Shared mailbox","Resource mailbox"]
    ol = [t for t in ordered if t in stats["type_counts"]]+[t for t in stats["type_counts"] if t not in ordered]
    os_ = [stats["type_counts"][t] for t in ol]
    ax=fig.add_axes([0.04,0.50,0.20,0.27]); _draw_donut(ax,os_,DONUT_COLORS,stats["total_mailboxes"],"Mailboxes")
    ax=fig.add_axes([0.25,0.54,0.20,0.22]); _draw_legend(ax,ol,os_,DONUT_COLORS)
    # Storage bar
    fig.text(0.54,0.80,"Mailboxes by storage used",fontsize=11,fontweight="bold",color=PRIMARY)
    ax=fig.add_axes([0.56,0.52,0.40,0.25])
    _draw_bar(ax,[b["label"] for b in stats["storage_buckets"]],[b["count"] for b in stats["storage_buckets"]])
    # Archive donut
    fig.text(0.04,0.42,"Archive status",fontsize=11,fontweight="bold",color=PRIMARY)
    ax=fig.add_axes([0.04,0.12,0.20,0.27]); _draw_donut(ax,[stats["archive_enabled"],stats["archive_disabled"]],DONUT_COLORS,stats["total_mailboxes"],"Mailboxes")
    ax=fig.add_axes([0.25,0.18,0.20,0.18]); _draw_legend(ax,["Enabled","Disabled"],[stats["archive_enabled"],stats["archive_disabled"]],DONUT_COLORS)
    # Top storage
    fig.text(0.54,0.42,"Top 10 by storage",fontsize=11,fontweight="bold",color=PRIMARY)
    ts=stats["top_storage"]; ax=fig.add_axes([0.63,0.10,0.33,0.28])
    _draw_hbar(ax,[r["email"].split("@")[0] for r in reversed(ts)],[r["storage"] for r in reversed(ts)],fmt=lambda v:f"{v:.1f} GB")
    pdf.savefig(fig,facecolor=fig.get_facecolor()); plt.close(fig)

    if detailed:
        fig=plt.figure(figsize=(14,9.5)); fig.patch.set_facecolor(BG)
        fig.text(0.04,0.96,"Exchange Online — Details",fontsize=17,fontweight="bold",color=TEXT,va="top")
        # Top items
        fig.text(0.04,0.92,"Top 10 by item count",fontsize=11,fontweight="bold",color=PRIMARY)
        ti=stats["top_items"]; ax=fig.add_axes([0.18,0.70,0.74,0.20])
        _draw_hbar(ax,[r["email"].split("@")[0] for r in reversed(ti)],[r["items"] for r in reversed(ti)],fmt=lambda v:f"{v:,}")
        # Table
        fig.text(0.04,0.64,"All Mailboxes",fontsize=11,fontweight="bold",color=PRIMARY)
        cols=["Email","Type","Items","Storage","Deleted","Del Size","Last Activity","Archive"]
        td=[[r["email"].split("@")[0],r["type"].replace(" mailbox",""),f'{r["items"]:,}',f'{r["storage"]:.2f}',f'{r["deleted"]:,}',f'{r["del_size"]:.2f}',r["last_activity"],r["archive"]] for r in stats["rows"]]
        ax=fig.add_axes([0.03,0.01,0.94,0.60]); _draw_table(ax,cols,td,[0.16,0.10,0.10,0.10,0.10,0.10,0.14,0.10])
        pdf.savefig(fig,facecolor=fig.get_facecolor()); plt.close(fig)

def _pdf_groups(pdf, stats, detailed):
    fig=plt.figure(figsize=(14,9.5)); fig.patch.set_facecolor(BG)
    fig.text(0.04,0.96,"Microsoft 365 Groups — Tenant Overview",fontsize=17,fontweight="bold",color=TEXT,va="top")
    sd=[(str(stats["total"]),"Total Groups"),(str(stats["no_owners"]),"Without Owners"),
        (str(stats["inactive"]),"Inactive Groups"),(str(stats["with_guests"]),"With Guests")]
    for i,(v,l) in enumerate(sd):
        ax=fig.add_axes([0.04+i*0.23,0.84,0.20,0.07]); _draw_stat(ax,v,l)
    # Privacy donut
    fig.text(0.04,0.78,"Group privacy distribution",fontsize=11,fontweight="bold",color=PRIMARY)
    pl=list(stats["privacy"].keys()); ps=list(stats["privacy"].values())
    ax=fig.add_axes([0.04,0.45,0.22,0.30]); _draw_donut(ax,ps,DONUT_COLORS,stats["total"],"Groups")
    ax=fig.add_axes([0.27,0.50,0.20,0.22]); _draw_legend(ax,pl,ps,DONUT_COLORS)
    # Storage bar
    fig.text(0.54,0.78,"SharePoint team sites by storage used",fontsize=11,fontweight="bold",color=PRIMARY)
    ax=fig.add_axes([0.56,0.48,0.40,0.27])
    _draw_bar(ax,[b["label"] for b in stats["storage_buckets"]],[b["count"] for b in stats["storage_buckets"]])
    pdf.savefig(fig,facecolor=fig.get_facecolor()); plt.close(fig)

    if detailed:
        fig=plt.figure(figsize=(14,9.5)); fig.patch.set_facecolor(BG)
        fig.text(0.04,0.96,"Microsoft 365 Groups — Details",fontsize=17,fontweight="bold",color=TEXT,va="top")
        cols=["Name","Privacy","Owners","Members","Guests","Teams","Last Activity","Total Size"]
        td=[[r["name"],r["privacy"],str(r["owners"]),str(r["members"]),str(r["guests"]),r["teams_status"],r["last_activity"],f'{r["total_size"]:.2f}'] for r in stats["rows"]]
        ax=fig.add_axes([0.03,0.05,0.94,0.85]); _draw_table(ax,cols,td)
        pdf.savefig(fig,facecolor=fig.get_facecolor()); plt.close(fig)

def _pdf_teams(pdf, stats, detailed):
    fig=plt.figure(figsize=(14,9.5)); fig.patch.set_facecolor(BG)
    fig.text(0.04,0.96,"Microsoft Teams — Tenant Overview",fontsize=17,fontweight="bold",color=TEXT,va="top")
    sd=[(str(stats["total"]),"Total Teams"),(str(stats["no_owners"]),"Without Owners"),
        (str(stats["inactive"]),"Inactive"),(str(stats["with_guests"]),"With Guests"),(str(stats["total_channels"]),"Total Channels")]
    for i,(v,l) in enumerate(sd):
        ax=fig.add_axes([0.04+i*0.185,0.84,0.17,0.07]); _draw_stat(ax,v,l)
    # Privacy donut
    fig.text(0.04,0.78,"Team privacy distribution",fontsize=11,fontweight="bold",color=PRIMARY)
    pl=list(stats["privacy"].keys()); ps=list(stats["privacy"].values())
    ax=fig.add_axes([0.04,0.45,0.22,0.30]); _draw_donut(ax,ps,DONUT_COLORS,stats["total"],"Teams")
    ax=fig.add_axes([0.27,0.50,0.20,0.22]); _draw_legend(ax,pl,ps,DONUT_COLORS)
    # Channel type donut
    fig.text(0.54,0.78,"Channel type distribution",fontsize=11,fontweight="bold",color=PRIMARY)
    cl=list(stats["channel_types"].keys()); cs=list(stats["channel_types"].values())
    ax=fig.add_axes([0.54,0.45,0.22,0.30]); _draw_donut(ax,cs,DONUT_COLORS,stats["total_channels"],"Channels")
    ax=fig.add_axes([0.77,0.50,0.20,0.22]); _draw_legend(ax,cl,cs,DONUT_COLORS)
    # Storage bar
    fig.text(0.04,0.38,"SharePoint team sites by storage used",fontsize=11,fontweight="bold",color=PRIMARY)
    ax=fig.add_axes([0.06,0.10,0.55,0.25])
    _draw_bar(ax,[b["label"] for b in stats["storage_buckets"]],[b["count"] for b in stats["storage_buckets"]])
    pdf.savefig(fig,facecolor=fig.get_facecolor()); plt.close(fig)

    if detailed:
        fig=plt.figure(figsize=(14,9.5)); fig.patch.set_facecolor(BG)
        fig.text(0.04,0.96,"Microsoft Teams — Details",fontsize=17,fontweight="bold",color=TEXT,va="top")
        cols=["Name","Privacy","Owners","Members","Guests","Channels","Status","Last Activity","Total Size"]
        td=[[r["name"],r["privacy"],str(r["owners"]),str(r["members"]),str(r["guests"]),str(r["total_channels"]),r["status"],r["last_activity"],f'{r["total_size"]:.2f}'] for r in stats["rows"]]
        ax=fig.add_axes([0.03,0.05,0.94,0.85]); _draw_table(ax,cols,td)
        pdf.savefig(fig,facecolor=fig.get_facecolor()); plt.close(fig)

def _pdf_onedrive(pdf, stats, detailed):
    fig=plt.figure(figsize=(14,9.5)); fig.patch.set_facecolor(BG)
    fig.text(0.04,0.96,"OneDrive — Tenant Overview",fontsize=17,fontweight="bold",color=TEXT,va="top")
    sd=[(str(stats["total"]),"Total Sites"),(f'{stats["total_storage"]:.2f}',"Storage (GB)"),(f'{stats["total_files"]:,}',"Total Files")]
    for i,(v,l) in enumerate(sd):
        ax=fig.add_axes([0.04+i*0.31,0.84,0.27,0.07]); _draw_stat(ax,v,l)
    # Storage bar
    fig.text(0.04,0.78,"Storage used",fontsize=11,fontweight="bold",color=PRIMARY)
    ax=fig.add_axes([0.06,0.50,0.40,0.25])
    _draw_bar(ax,[b["label"] for b in stats["storage_buckets"]],[b["count"] for b in stats["storage_buckets"]])
    # Top files
    fig.text(0.54,0.78,"Top 10 by file count",fontsize=11,fontweight="bold",color=PRIMARY)
    tf=stats["top_files"]; ax=fig.add_axes([0.63,0.48,0.33,0.27])
    _draw_hbar(ax,[r["display_name"] for r in reversed(tf)],[r["files"] for r in reversed(tf)],color=TEAL,fmt=lambda v:f"{v:,}")
    pdf.savefig(fig,facecolor=fig.get_facecolor()); plt.close(fig)

    if detailed:
        fig=plt.figure(figsize=(14,9.5)); fig.patch.set_facecolor(BG)
        fig.text(0.04,0.96,"OneDrive — Details",fontsize=17,fontweight="bold",color=TEXT,va="top")
        cols=["User","Storage (GB)","Files","Last Activity"]
        td=[[r["display_name"],f'{r["storage"]:.2f}',f'{r["files"]:,}',r["last_activity"]] for r in stats["rows"]]
        ax=fig.add_axes([0.08,0.05,0.84,0.85]); _draw_table(ax,cols,td,[0.30,0.20,0.20,0.20])
        pdf.savefig(fig,facecolor=fig.get_facecolor()); plt.close(fig)

def _pdf_sharepoint(pdf, stats, detailed):
    fig=plt.figure(figsize=(14,9.5)); fig.patch.set_facecolor(BG)
    fig.text(0.04,0.96,"SharePoint Online — Tenant Overview",fontsize=17,fontweight="bold",color=TEXT,va="top")
    sd=[(str(stats["total"]),"Total Sites"),(f'{stats["total_storage"]:.2f}',"Storage (GB)"),(f'{stats["total_files"]:,}',"Total Files")]
    for i,(v,l) in enumerate(sd):
        ax=fig.add_axes([0.04+i*0.31,0.84,0.27,0.07]); _draw_stat(ax,v,l)
    # Storage bar
    fig.text(0.04,0.78,"Site collections by storage used",fontsize=11,fontweight="bold",color=PRIMARY)
    ax=fig.add_axes([0.06,0.50,0.40,0.25])
    _draw_bar(ax,[b["label"] for b in stats["storage_buckets"]],[b["count"] for b in stats["storage_buckets"]])
    # Activity bar
    fig.text(0.54,0.78,"Site collections by last activity",fontsize=11,fontweight="bold",color=PRIMARY)
    ax=fig.add_axes([0.56,0.50,0.40,0.25])
    _draw_bar(ax,list(stats["activity"].keys()),list(stats["activity"].values()))
    # Templates
    fig.text(0.04,0.40,"Top templates by site count",fontsize=11,fontweight="bold",color=PRIMARY)
    tl=sorted(stats["templates"].items(),key=lambda x:x[1],reverse=True)[:10]
    ax=fig.add_axes([0.14,0.10,0.30,0.27])
    _draw_hbar(ax,[t[0][:20] for t in reversed(tl)],[t[1] for t in reversed(tl)],color=TEAL)
    # Top files
    fig.text(0.54,0.40,"Top 10 by file count",fontsize=11,fontweight="bold",color=PRIMARY)
    tf=stats["top_files"]; ax=fig.add_axes([0.63,0.10,0.33,0.27])
    _draw_hbar(ax,[r["name"][:18] for r in reversed(tf)],[r["files"] for r in reversed(tf)],color=PRIMARY,fmt=lambda v:f"{v:,}")
    pdf.savefig(fig,facecolor=fig.get_facecolor()); plt.close(fig)

    if detailed:
        fig=plt.figure(figsize=(14,9.5)); fig.patch.set_facecolor(BG)
        fig.text(0.04,0.96,"SharePoint Online — Details",fontsize=17,fontweight="bold",color=TEXT,va="top")
        cols=["Site Name","Template","Storage (GB)","Files","Last Activity"]
        td=[[r["name"],r["template"],f'{r["storage"]:.2f}',f'{r["files"]:,}',r["last_activity"]] for r in stats["rows"]]
        ax=fig.add_axes([0.05,0.05,0.90,0.85]); _draw_table(ax,cols,td,[0.25,0.22,0.15,0.15,0.18])
        pdf.savefig(fig,facecolor=fig.get_facecolor()); plt.close(fig)


def generate_pdf(all_stats, detailed=True):
    _setup_plt()
    buf = io.BytesIO()
    with PdfPages(buf) as pdf:
        if "exchange" in all_stats: _pdf_exchange(pdf, all_stats["exchange"], detailed)
        if "groups" in all_stats: _pdf_groups(pdf, all_stats["groups"], detailed)
        if "teams" in all_stats: _pdf_teams(pdf, all_stats["teams"], detailed)
        if "onedrive" in all_stats: _pdf_onedrive(pdf, all_stats["onedrive"], detailed)
        if "sharepoint" in all_stats: _pdf_sharepoint(pdf, all_stats["sharepoint"], detailed)
    buf.seek(0)
    return buf


# ─── Routes ──────────────────────────────────────────────────────────────────

WORKLOADS = {
    "exchange": {"label": "Exchange Online", "parser": parse_exchange, "stats_fn": compute_exchange_stats,
                 "desc": "Mailbox details — items, storage, types, archive status"},
    "groups": {"label": "Microsoft 365 Groups", "parser": parse_groups, "stats_fn": compute_groups_stats,
               "desc": "Group details — privacy, owners, members, guests, Teams status"},
    "teams": {"label": "Microsoft Teams", "parser": parse_teams, "stats_fn": compute_teams_stats,
              "desc": "Team details — channels, members, privacy, storage"},
    "onedrive": {"label": "OneDrive", "parser": parse_onedrive, "stats_fn": compute_onedrive_stats,
                 "desc": "OneDrive site details — user storage, files, activity"},
    "sharepoint": {"label": "SharePoint Online", "parser": parse_sharepoint, "stats_fn": compute_sharepoint_stats,
                   "desc": "Site collection details — templates, storage, files"},
}

@app.route("/")
def index():
    return render_template("index.html", workloads=WORKLOADS)

@app.route("/upload", methods=["POST"])
def upload():
    all_stats = {}
    uploaded_any = False

    for wk, winfo in WORKLOADS.items():
        f = request.files.get(f"file_{wk}")
        if f and f.filename:
            uploaded_any = True
            uid = uuid.uuid4().hex[:8]
            path = os.path.join(app.config["UPLOAD_FOLDER"], f"{uid}_{f.filename}")
            f.save(path)
            try:
                rows = winfo["parser"](path)
                if rows:
                    all_stats[wk] = winfo["stats_fn"](rows)
            except Exception as e:
                flash(f'Error parsing {winfo["label"]}: {e}', "error")

    if not uploaded_any:
        flash("Please upload at least one file.", "error")
        return redirect(url_for("index"))

    if not all_stats:
        flash("No valid data found in the uploaded files.", "error")
        return redirect(url_for("index"))

    # Store stats to a JSON file (too large for cookie session)
    data_id = uuid.uuid4().hex[:12]
    data_path = os.path.join(app.config["UPLOAD_FOLDER"], f"stats_{data_id}.json")
    with open(data_path, "w") as fp:
        json.dump({"stats": all_stats, "labels": {k: WORKLOADS[k]["label"] for k in all_stats}}, fp, default=str)
    session["data_id"] = data_id
    return redirect(url_for("dashboard"))


def _load_stats():
    """Load stats from the JSON file referenced in the session."""
    data_id = session.get("data_id")
    if not data_id:
        return None, None
    data_path = os.path.join(app.config["UPLOAD_FOLDER"], f"stats_{data_id}.json")
    if not os.path.exists(data_path):
        return None, None
    with open(data_path) as fp:
        data = json.load(fp)
    return data["stats"], data["labels"]


@app.route("/dashboard")
def dashboard():
    all_stats, labels = _load_stats()
    if not all_stats:
        flash("Please upload files first.", "error")
        return redirect(url_for("index"))
    now_str = datetime.now().strftime("%b %d, %Y at %I:%M %p")
    return render_template("dashboard.html", all_stats=all_stats, labels=labels, now=now_str)

@app.route("/export-pdf/<mode>")
def export_pdf(mode):
    all_stats, _ = _load_stats()
    if not all_stats:
        flash("No data to export.", "error")
        return redirect(url_for("index"))
    detailed = mode == "detailed"
    buf = generate_pdf(all_stats, detailed=detailed)
    suffix = "Detailed" if detailed else "Summary"
    return send_file(buf, mimetype="application/pdf", as_attachment=True,
                     download_name=f"M365_Tenant_Report_{suffix}.pdf")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
